"""
ClipControl Backend - API Principal
"""

from fastapi import FastAPI, HTTPException, Form, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from typing import List, Optional
from datetime import datetime, date, timedelta
import bcrypt
import hashlib
import secrets
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse
import io

from config import settings
from database import get_supabase
from models import (
    LoginRequest, LoginResponse, MessageResponse,
    EmpleadoCreate, EmpleadoUpdate,
    EntregaCreate,
    ValidarRetiroRequest, ValidarRetiroResponse,
    PeriodoCreate,
    UsuarioCreate, UsuarioUpdate,
    GenerarQRMasivoRequest,
)

# Obtener cliente de Supabase
supabase = get_supabase()


# Inicializar FastAPI
app = FastAPI(
    title=settings.APP_NAME,
    version=settings.APP_VERSION,
    debug=settings.DEBUG
)

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# RUTAS SALUD
# ==========================================

@app.get("/")
def read_root():
    return {
        "app": settings.APP_NAME,
        "version": settings.APP_VERSION,
        "status": "running",
        "message": "API de ClipControl funcionando correctamente"
    }


@app.get("/health")
def health_check():
    try:
        supabase = get_supabase()
        supabase.table("sucursales").select("count", count="exact").execute()
        return {
            "status": "healthy",
            "database": "connected",
            "timestamp": str(datetime.now())
        }
    except Exception as e:
        return JSONResponse(
            status_code=503,
            content={"status": "unhealthy", "error": str(e)}
        )

# ==========================================
# LOGIN
# ==========================================

@app.post("/api/auth/login", response_model=LoginResponse)
def login(request: LoginRequest):
    """Login de usuario"""
    try:
        supabase = get_supabase()
        
        # Buscar usuario
        result = supabase.table("usuarios").select("*").eq("username", request.username).execute()
        
        if not result.data or len(result.data) == 0:
            raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
        
        user = result.data[0]
        
        # Verificar contraseña
        password_match = bcrypt.checkpw(
            request.password.encode('utf-8'),
            user['password_hash'].encode('utf-8')
        )
        
        if not password_match:
            raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
        
        # Verificar que el usuario esté activo
        if not user.get('activo', True):
            raise HTTPException(status_code=401, detail="Usuario inactivo")
        
        # Actualizar último acceso
        supabase.table("usuarios").update({
            "ultimo_acceso": datetime.now().isoformat()
        }).eq("id", user['id']).execute()
        
        # Generar token
        token = f"token_{user['id']}_{user['username']}"
        
        # Preparar respuesta
        user_data = {
            "id": user['id'],
            "username": user['username'],
            "rol": user['rol'].lower() if user['rol'] else None,
            "nombre_completo": user.get('nombre_completo'),
            "sucursal_id": user.get('sucursal_id')
        }
        
        return LoginResponse(
            access_token=token,
            token_type="bearer",
            user=user_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en login: {str(e)}")

# ==========================================
# EMPLEADOS
# ==========================================

@app.get("/api/empleados", response_model=List[dict])
def get_empleados(
    sucursal_id: Optional[int] = None,
    tipo_contrato: Optional[str] = None,
    activo: Optional[bool] = True,
    skip: int = 0,
    limit: int = 100
):
    try:
        supabase = get_supabase()
        query = supabase.table("v_empleados_completo").select("*")

        if activo is not None:
            query = query.eq("activo", activo)
        if sucursal_id:
            query = query.eq("sucursal_id", sucursal_id)
        if tipo_contrato:
            query = query.eq("tipo_contrato", tipo_contrato)

        result = query.range(skip, skip + limit - 1).execute()
        return result.data

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener empleados: {str(e)}")


@app.get("/api/empleados/{empleado_id}", response_model=dict)
def get_empleado(empleado_id: int):
    try:
        supabase = get_supabase()
        result = supabase.table("empleados").select("*").eq("id", empleado_id).execute()

        if not result.data:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")

        return result.data[0]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/empleados/rut/{rut}", response_model=dict)
def get_empleado_by_rut(rut: str):
    try:
        supabase = get_supabase()
        rut_limpio = rut.replace(".", "").replace("-", "")

        result = supabase.table("v_empleados_completo").select("*").eq("rut", rut_limpio).execute()

        if not result.data:
            raise HTTPException(status_code=404, detail=f"Empleado con RUT {rut} no encontrado")

        return result.data[0]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/api/empleados", response_model=dict)
def create_empleado(empleado: EmpleadoCreate):
    try:
        supabase = get_supabase()
        existing = supabase.table("empleados").select("id").eq("rut", empleado.rut).execute()

        if existing.data:
            raise HTTPException(status_code=400, detail="Ya existe un empleado con ese RUT")

        result = supabase.table("empleados").insert(empleado.dict()).execute()
        return result.data[0]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.put("/api/empleados/{empleado_id}", response_model=dict)
def update_empleado(empleado_id: int, empleado: EmpleadoUpdate):
    try:
        supabase = get_supabase()
        existing = supabase.table("empleados").select("id").eq("id", empleado_id).execute()

        if not existing.data:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")

        update_data = {k: v for k, v in empleado.dict().items() if v is not None}
        result = supabase.table("empleados").update(update_data).eq("id", empleado_id).execute()

        return result.data[0]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# VALIDACIÓN RETIRO
# ==========================================

@app.post("/api/validar-retiro", response_model=ValidarRetiroResponse)
def validar_retiro(validacion: ValidarRetiroRequest):
    try:
        supabase = get_supabase()
        rut_limpio = validacion.rut.replace(".", "").replace("-", "")

        # Buscar empleado
        empleado_result = supabase.table("empleados").select("*").eq("rut", rut_limpio).eq("activo", True).execute()

        if not empleado_result.data:
            return ValidarRetiroResponse(
                valido=False,
                mensaje=f"Empleado con RUT {validacion.rut} no encontrado o inactivo",
                ya_retiro=False
            )

        empleado = empleado_result.data[0]

        # Revisar si ya retiró
        entrega_result = supabase.table("entregas").select("id, fecha_hora").eq(
            "empleado_id", empleado['id']
        ).eq(
            "periodo_id", validacion.periodo_id
        ).eq("estado", "COMPLETADO").execute()

        if entrega_result.data:
            entrega = entrega_result.data[0]
            return ValidarRetiroResponse(
                valido=False,
                mensaje=f"Este empleado ya retiró el {entrega['fecha_hora']}",
                empleado=empleado,
                ya_retiro=True,
                tipo_caja=None
            )

        tipo_caja = "PLANTA" if empleado['tipo_contrato'] == "PLANTA" else "PLAZO_FIJO"

        return ValidarRetiroResponse(
            valido=True,
            mensaje="Empleado válido para retiro",
            empleado=empleado,
            ya_retiro=False,
            tipo_caja=tipo_caja
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en validación: {str(e)}")

@app.get("/api/entregas/lista")
def get_entregas_lista(
    periodo_id: Optional[int] = None, 
    sucursal_id: Optional[int] = None,
    tipo_contrato: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    skip: int = 0, 
    limit: int = 200
):
    """Obtener lista de entregas con filtros combinados y rango de fechas"""
    try:
        supabase = get_supabase()
        
        # Consulta base
        query = supabase.table("entregas").select(
            "*, empleados!inner(*, sucursales(*))"
        )

        # SI HAY FILTRO DE FECHAS, NO FILTRAR POR PERÍODO
        if fecha_desde or fecha_hasta:
            # Solo aplicar filtros de fecha
            if fecha_desde:
                query = query.gte("fecha_hora", f"{fecha_desde}T00:00:00")
            if fecha_hasta:
                query = query.lte("fecha_hora", f"{fecha_hasta}T23:59:59")
        else:
            # Si NO hay fechas, filtrar por período activo
            if periodo_id:
                query = query.eq("periodo_id", periodo_id)
            else:
                periodo_activo = supabase.table("periodos_entrega").select("id").eq("activo", True).execute()
                if periodo_activo.data:
                    query = query.eq("periodo_id", periodo_activo.data[0]['id'])

        # Filtros de sucursal y tipo
        if sucursal_id:
            query = query.eq("empleados.sucursal_id", sucursal_id)
        
        if tipo_contrato:
            query = query.eq("empleados.tipo_contrato", tipo_contrato)

        # Ejecutar query
        result = query.order("fecha_hora", desc=True).range(skip, skip + limit - 1).execute()
        
        print(f"📊 Filtros: desde={fecha_desde}, hasta={fecha_hasta}, sucursal={sucursal_id}, tipo={tipo_contrato}, Resultados={len(result.data)}")
        
        return result.data

    except Exception as e:
        print(f"❌ Error en backend: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al obtener entregas: {str(e)}")


@app.put("/api/entregas/{entrega_id}")
def update_entrega(entrega_id: int, datos: dict):
    """Actualizar observaciones de una entrega"""
    try:
        supabase = get_supabase()
        
        # Verificar que la entrega existe
        entrega = supabase.table("entregas").select("id").eq("id", entrega_id).execute()
        if not entrega.data:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")
        
        # Actualizar solo observaciones
        update_data = {}
        if "observaciones" in datos:
            update_data["observaciones"] = datos["observaciones"]
        
        result = supabase.table("entregas").update(update_data).eq("id", entrega_id).execute()
        
        return {"message": "Entrega actualizada correctamente", "data": result.data[0]}
        
    except Exception as e:
        print(f"❌ ERROR en update_entrega: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.delete("/api/entregas/{entrega_id}")
def delete_entrega(entrega_id: int):
    """Cancelar/eliminar una entrega"""
    try:
        supabase = get_supabase()
        
        # Verificar que la entrega existe
        entrega = supabase.table("entregas").select("id, qr_token_id").eq("id", entrega_id).execute()
        if not entrega.data:
            raise HTTPException(status_code=404, detail="Entrega no encontrada")
        
        entrega_data = entrega.data[0]
        
        # Eliminar la entrega
        supabase.table("entregas").delete().eq("id", entrega_id).execute()
        
        # Opcional: Reactivar el token QR si existe
        if entrega_data.get("qr_token_id"):
            supabase.table("qr_tokens").update({"usado": False}).eq("id", entrega_data["qr_token_id"]).execute()
        
        return {"message": "Entrega cancelada correctamente"}
        
    except Exception as e:
        print(f"❌ ERROR en delete_entrega: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/entregas/estadisticas-guardia/{usuario_id}")
def get_estadisticas_guardia(usuario_id: int):
    """
    Obtener estadísticas de entregas del día de un guardia específico
    """
    try:
        supabase = get_supabase()
        
        # Obtener fecha de hoy
        hoy = date.today().isoformat()
        
        # Entregas de hoy de este guardia
        entregas_hoy = supabase.table("entregas").select("*, empleados(tipo_contrato)").eq(
            "usuario_id", usuario_id
        ).gte(
            "fecha_hora", f"{hoy}T00:00:00"
        ).execute()
        
        # Contar por tipo
        total = len(entregas_hoy.data) if entregas_hoy.data else 0
        planta = sum(1 for e in (entregas_hoy.data or []) if e.get('empleados', {}).get('tipo_contrato') == 'PLANTA')
        plazo_fijo = total - planta
        
        # Última entrega
        ultima_entrega = None
        if entregas_hoy.data and len(entregas_hoy.data) > 0:
            # Ordenar por fecha_hora descendente
            entregas_ordenadas = sorted(entregas_hoy.data, key=lambda x: x.get('fecha_hora', ''), reverse=True)
            if entregas_ordenadas:
                ultima_fecha = entregas_ordenadas[0].get('fecha_hora')
                if ultima_fecha:
                    ultima_dt = datetime.fromisoformat(ultima_fecha.replace('Z', '+00:00'))
                    ahora = datetime.now(ultima_dt.tzinfo)
                    diferencia = int((ahora - ultima_dt).total_seconds() / 60)  # minutos
                    ultima_entrega = {
                        "fecha": ultima_fecha,
                        "minutos_atras": diferencia
                    }
        
        return {
            "total_hoy": total,
            "planta": planta,
            "plazo_fijo": plazo_fijo,
            "ultima_entrega": ultima_entrega
        }
        
    except Exception as e:
        print(f"❌ ERROR en get_estadisticas_guardia: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/api/entregas/estadisticas")
def get_estadisticas_entregas():
    """Obtener estadísticas completas de entregas"""
    try:
        supabase = get_supabase()
        
        # Total de entregas
        total = supabase.table("entregas").select("id", count="exact").execute()
        
        # Entregas de hoy
        hoy = date.today().isoformat()
        hoy_entregas = supabase.table("entregas").select("id", count="exact").gte(
            "fecha_hora", f"{hoy}T00:00:00"
        ).execute()
        
        # Entregas por tipo de contrato (de hoy)
        entregas_hoy_detalle = supabase.table("entregas").select(
            "id, empleados(tipo_contrato)"
        ).gte("fecha_hora", f"{hoy}T00:00:00").execute()
        
        planta = sum(1 for e in (entregas_hoy_detalle.data or []) 
                    if e.get('empleados', {}).get('tipo_contrato') == 'PLANTA')
        plazo_fijo = len(entregas_hoy_detalle.data or []) - planta
        
        return {
            "total": total.count if total.count else 0,
            "hoy": hoy_entregas.count if hoy_entregas.count else 0,
            "planta": planta,
            "plazo_fijo": plazo_fijo
        }
    except Exception as e:
        print(f"❌ ERROR en get_estadisticas_entregas: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
@app.get("/api/dashboard/estadisticas")
def get_estadisticas_dashboard():
    """Obtener estadísticas generales del dashboard principal"""
    try:
        supabase = get_supabase()
        
        # Total empleados activos
        empleados = supabase.table("empleados").select("id", count="exact").eq("activo", True).execute()
        total_empleados = empleados.count if empleados.count else 0
        
        # Entregas de hoy
        hoy = date.today().isoformat()
        entregas_hoy = supabase.table("entregas").select("id", count="exact").gte(
            "fecha_hora", f"{hoy}T00:00:00"
        ).execute()
        total_hoy = entregas_hoy.count if entregas_hoy.count else 0
        
        # Porcentaje de entregas (empleados que ya retiraron hoy)
        porcentaje = round((total_hoy / total_empleados * 100), 1) if total_empleados > 0 else 0
        
        # Pendientes (empleados que no han retirado hoy)
        pendientes = total_empleados - total_hoy
        
        return {
            "total_empleados": total_empleados,
            "entregas_hoy": total_hoy,
            "porcentaje": porcentaje,
            "pendientes": pendientes
        }
        
    except Exception as e:
        print(f"❌ ERROR en get_estadisticas_dashboard: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/api/entregas", response_model=dict)
def crear_entrega(entrega: EntregaCreate):
    """Registrar una nueva entrega"""
    try:
        supabase = get_supabase()

        # Verificar si ya retiró
        if entrega.periodo_id:
            check = supabase.table("entregas").select("id").eq(
                "empleado_id", entrega.empleado_id
            ).eq(
                "periodo_id", entrega.periodo_id
            ).eq(
                "estado", "COMPLETADO"
            ).execute()

            if check.data:
                raise HTTPException(status_code=400, detail="Este empleado ya retiró en este período")

        # Crear diccionario con todos los campos
        data = entrega.dict(exclude_none=True)
        
        # Asegurarse de que tenga valores por defecto
        if "metodo" not in data:
            data["metodo"] = "QR"
        if "estado" not in data:
            data["estado"] = "COMPLETADO"

        result = supabase.table("entregas").insert(data).execute()
        return result.data[0]

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear entrega: {str(e)}")

# ==========================================
# PERIODOS
# ==========================================

@app.get("/api/periodos", response_model=List[dict])
def get_periodos(activo: Optional[bool] = None):
    try:
        supabase = get_supabase()
        query = supabase.table("periodos_entrega").select("*")

        if activo is not None:
            query = query.eq("activo", activo)

        result = query.order("fecha_inicio", desc=True).execute()
        return result.data

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/periodos/activo", response_model=dict)
def get_periodo_activo():
    try:
        supabase = get_supabase()
        result = supabase.table("periodos_entrega").select("*").eq("activo", True).execute()

        if not result.data:
            raise HTTPException(status_code=404, detail="No hay período activo")

        return result.data[0]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ==========================================
# ==========================================
# SUCURSALES
# ==========================================

@app.get("/api/sucursales", response_model=List[dict])
def get_sucursales(activa: Optional[bool] = None):
    try:
        supabase = get_supabase()
        query = supabase.table("sucursales").select("*")
        
        if activa is not None:
            query = query.eq("activa", activa)
        
        result = query.order("nombre").execute()
        return result.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/api/sucursales")
async def create_sucursal(sucursal: dict):
    try:
        supabase = get_supabase()
        
        # Verificar que no exista
        check = supabase.table("sucursales").select("id").eq("nombre", sucursal['nombre']).execute()
        if check.data:
            raise HTTPException(status_code=400, detail="Ya existe una sucursal con ese nombre")
        
        data = {
            "nombre": sucursal['nombre'],
            "direccion": sucursal.get('direccion'),
            "ciudad": sucursal.get('ciudad'),
            "activa": sucursal.get('activa', True)
        }
        
        result = supabase.table("sucursales").insert(data).execute()
        return result.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/sucursales/{sucursal_id}")
async def update_sucursal(sucursal_id: int, sucursal: dict):
    try:
        supabase = get_supabase()
        
        data = {}
        if 'nombre' in sucursal:
            data['nombre'] = sucursal['nombre']
        if 'direccion' in sucursal:
            data['direccion'] = sucursal['direccion']
        if 'ciudad' in sucursal:
            data['ciudad'] = sucursal['ciudad']
        if 'activa' in sucursal:
            data['activa'] = sucursal['activa']
        
        result = supabase.table("sucursales").update(data).eq("id", sucursal_id).execute()
        return {"message": "Sucursal actualizada correctamente", "data": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/sucursales/{sucursal_id}")
async def delete_sucursal(sucursal_id: int):
    try:
        supabase = get_supabase()
        
        # Verificar que no tenga empleados
        empleados = supabase.table("empleados").select("id", count="exact").eq("sucursal_id", sucursal_id).execute()
        if empleados.count and empleados.count > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"No puedes eliminar esta sucursal porque tiene {empleados.count} empleados asignados"
            )
        
        # Desactivar en vez de eliminar
        result = supabase.table("sucursales").update({"activa": False}).eq("id", sucursal_id).execute()
        return {"message": "Sucursal desactivada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# ==========================================
# REPORTES
# ==========================================

@app.get("/api/reportes/entregas-por-sucursal")
def get_entregas_por_sucursal():
    """Reporte de entregas agrupadas por sucursal"""
    try:
        supabase = get_supabase()
        result = supabase.table("entregas") \
            .select("id, empleados(sucursal_id, sucursales(nombre))") \
            .execute()
        
        # Agrupar por sucursal
        sucursales = {}
        for entrega in result.data:
            if entrega.get("empleados") and entrega["empleados"].get("sucursales"):
                sucursal = entrega["empleados"]["sucursales"]["nombre"]
                sucursales[sucursal] = sucursales.get(sucursal, 0) + 1
        
        return [{"sucursal": k, "total": v} for k, v in sucursales.items()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/reportes/entregas-por-fecha")
def get_entregas_por_fecha(
    fecha_inicio: str = None, 
    fecha_fin: str = None,
    sucursal_id: Optional[int] = None,
    tipo_contrato: Optional[str] = None
):
    """Reporte de entregas filtradas por fecha, sucursal y tipo de contrato"""
    try:
        supabase = get_supabase()
        
        # Obtener todas las entregas con filtros de fecha
        query = supabase.table("entregas").select(
            "*, empleados(id, rut, nombre, apellido, tipo_contrato, sucursal_id, sucursales(nombre))"
        )
        
        if fecha_inicio:
            query = query.gte("fecha_hora", f"{fecha_inicio}T00:00:00")
        if fecha_fin:
            query = query.lte("fecha_hora", f"{fecha_fin}T23:59:59")
        
        result = query.order("fecha_hora", desc=True).execute()
        
        # Filtrar en Python (porque Supabase no soporta filtros en relaciones anidadas)
        entregas = result.data
        
        if sucursal_id:
            entregas = [e for e in entregas if e.get('empleados', {}).get('sucursal_id') == sucursal_id]
        
        if tipo_contrato:
            entregas = [e for e in entregas if e.get('empleados', {}).get('tipo_contrato') == tipo_contrato]
        
        # Adaptar respuesta
        for entrega in entregas:
            if 'fecha_hora' in entrega:
                entrega['fecha_retiro'] = entrega['fecha_hora']
        
        print(f"📊 Filtros aplicados: sucursal_id={sucursal_id}, tipo_contrato={tipo_contrato}, Total={len(entregas)}")
        
        return entregas
        
    except Exception as e:
        print(f"❌ ERROR en get_entregas_por_fecha: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/reportes/resumen")
def get_resumen_general():
    """Resumen general del sistema"""
    try:
        supabase = get_supabase()
        
        # Total empleados
        total_empleados = supabase.table("empleados").select("id", count="exact").execute()
        empleados_activos = supabase.table("empleados").select("id", count="exact").eq("activo", True).execute()
        
        # Total entregas
        total_entregas = supabase.table("entregas").select("id", count="exact").execute()
        
        # Sucursales
        sucursales = supabase.table("sucursales").select("*").eq("activa", True).execute()
        
        return {
            "empleados": {
                "total": total_empleados.count if total_empleados.count else 0,
                "activos": empleados_activos.count if empleados_activos.count else 0,
                "inactivos": (total_empleados.count - empleados_activos.count) if total_empleados.count else 0
            },
            "entregas": {
                "total": total_entregas.count if total_entregas.count else 0
            },
            "sucursales": len(sucursales.data) if sucursales.data else 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# CONFIGURACIÓN DEL SISTEMA (SUPERADMIN)
# ==========================================

@app.get("/api/configuracion")
def get_configuracion():
    """Obtener configuración del sistema"""
    try:
        supabase = get_supabase()
        result = supabase.table("configuracion_sistema").select("*").execute()
        
        # Convertir a diccionario
        config = {}
        for item in result.data:
            config[item['clave']] = item['valor']
        
        return config
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/api/configuracion")
def update_configuracion(config: dict):
    """Actualizar configuración del sistema (solo SUPERADMIN)"""
    try:
        supabase = get_supabase()
        
        # Aquí deberías validar que el usuario es SUPERADMIN
        # Por ahora lo dejamos sin validación
        
        for clave, valor in config.items():
            supabase.table("configuracion_sistema").update({
                "valor": str(valor),
                "updated_at": datetime.now().isoformat()
            }).eq("clave", clave).execute()
        
        return {"message": "Configuración actualizada exitosamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# BENEFICIOS
# ==========================================

@app.get("/api/beneficios")
def get_beneficios(activo: Optional[bool] = None):
    """Obtener tipos de beneficios"""
    try:
        supabase = get_supabase()
        query = supabase.table("tipos_beneficio").select("*")
        
        if activo is not None:
            query = query.eq("activo", activo)
        
        result = query.order("nombre").execute()
        return result.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/api/beneficios")
def create_beneficio(
    nombre: str,
    descripcion: Optional[str] = None,
    frecuencia: str = "EVENTUAL",
    mes_entrega: Optional[int] = None
):
    """Crear nuevo tipo de beneficio"""
    try:
        supabase = get_supabase()
        data = {
            "nombre": nombre,
            "descripcion": descripcion,
            "frecuencia": frecuencia,
            "mes_entrega": mes_entrega,
            "activo": True
        }
        result = supabase.table("tipos_beneficio").insert(data).execute()
        return result.data[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# CAJAS NO RETIRADAS
# ==========================================

@app.get("/api/cajas-no-retiradas")
def get_cajas_no_retiradas(estado: Optional[str] = None):
    """Obtener cajas no retiradas"""
    try:
        supabase = get_supabase()
        query = supabase.table("v_cajas_no_retiradas").select("*")
        
        if estado:
            query = query.eq("estado", estado)
        
        result = query.order("fecha_limite", desc=True).execute()
        return result.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/api/verificar-pendientes/{periodo_id}")
def verificar_pendientes_periodo(periodo_id: int):
    """
    Verificar empleados que no retiraron y marcar cajas para sindicato
    """
    try:
        supabase = get_supabase()
        
        # Obtener período
        periodo = supabase.table("periodos_entrega").select("*").eq("id", periodo_id).execute()
        if not periodo.data:
            raise HTTPException(status_code=404, detail="Período no encontrado")
        
        periodo_data = periodo.data[0]
        fecha_fin = periodo_data['fecha_fin']
        
        # Obtener todos los empleados activos
        empleados = supabase.table("empleados").select("id").eq("activo", True).execute()
        empleados_ids = [e['id'] for e in empleados.data]
        
        # Obtener empleados que SÍ retiraron
        entregas = supabase.table("entregas").select("empleado_id").eq("periodo_id", periodo_id).execute()
        retiraron_ids = [e['empleado_id'] for e in entregas.data]
        
        # Empleados que NO retiraron
        no_retiraron_ids = [e_id for e_id in empleados_ids if e_id not in retiraron_ids]
        
        # Marcar cajas como no retiradas
        for empleado_id in no_retiraron_ids:
            data = {
                "periodo_id": periodo_id,
                "empleado_id": empleado_id,
                "tipo_beneficio_id": periodo_data.get('tipo_beneficio_id', 1),
                "fecha_limite": fecha_fin,
                "estado": "TRASPASADO_SINDICATO",
                "fecha_traspaso": datetime.now().isoformat(),
                "observaciones": f"No retiró en plazo. Traspasado automáticamente al sindicato."
            }
            supabase.table("cajas_no_retiradas").insert(data).execute()
        
        return {
            "total_empleados": len(empleados_ids),
            "retiraron": len(retiraron_ids),
            "no_retiraron": len(no_retiraron_ids),
            "traspasados_sindicato": len(no_retiraron_ids)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# ACTUALIZAR ENDPOINT DE USUARIOS
# ==========================================

# Reemplaza tu endpoint POST /api/usuarios con este:

@app.post("/api/usuarios", response_model=dict)
def create_usuario(usuario: UsuarioCreate):
    """Crear nuevo usuario/guardia (con validación de roles)"""
    try:
        supabase = get_supabase()
        
        # IMPORTANTE: En producción, validar el rol del usuario que hace la petición
        # Por ejemplo, un ADMIN solo puede crear GUARDIAS
        
        # Validar que ADMIN solo puede crear GUARDIAS
        # if usuario_actual.rol == 'ADMIN' and usuario.rol != 'GUARDIA':
        #     raise HTTPException(status_code=403, detail="Solo puedes crear usuarios GUARDIA")
        
        # Verificar si el username ya existe
        existing = supabase.table("usuarios").select("id").eq("username", usuario.username).execute()
        if existing.data:
            raise HTTPException(status_code=400, detail="El nombre de usuario ya existe")
        
        # Hashear la contraseña
        password_hash = bcrypt.hashpw(usuario.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Crear usuario
        data = {
            "username": usuario.username,
            "password_hash": password_hash,
            "rol": usuario.rol.upper(),
            "nombre_completo": usuario.nombre_completo,
            "sucursal_id": usuario.sucursal_id,
            "activo": usuario.activo
        }
        
        result = supabase.table("usuarios").insert(data).execute()
        
        # No devolver el password_hash
        user = result.data[0]
        user.pop('password_hash', None)
        return user
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# GESTIÓN DE USUARIOS (GUARDIAS)
# ==========================================

@app.get("/api/usuarios")
def get_usuarios(rol: Optional[str] = None):
    """Obtener lista de usuarios"""
    try:
        supabase = get_supabase()
        query = supabase.table("usuarios").select("id, username, rol, nombre_completo, sucursal_id, activo, ultimo_acceso")
        
        if rol:
            query = query.eq("rol", rol.upper())
        
        result = query.order("nombre_completo").execute()
        return result.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/api/usuarios/{usuario_id}")
def update_usuario(usuario_id: int, usuario: UsuarioUpdate):
    """Actualizar usuario/guardia"""
    try:
        supabase = get_supabase()
        
        # Verificar que existe
        existing = supabase.table("usuarios").select("id").eq("id", usuario_id).execute()
        if not existing.data:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        # Preparar datos
        update_data = {}
        if usuario.username:
            update_data["username"] = usuario.username
        if usuario.password:
            update_data["password_hash"] = bcrypt.hashpw(usuario.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        if usuario.nombre_completo:
            update_data["nombre_completo"] = usuario.nombre_completo
        if usuario.sucursal_id is not None:
            update_data["sucursal_id"] = usuario.sucursal_id
        if usuario.activo is not None:
            update_data["activo"] = usuario.activo
        
        result = supabase.table("usuarios").update(update_data).eq("id", usuario_id).execute()
        
        user = result.data[0]
        user.pop('password_hash', None)
        return user
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/api/usuarios/{usuario_id}")
def delete_usuario(usuario_id: int):
    """Desactivar usuario (soft delete)"""
    try:
        supabase = get_supabase()
        result = supabase.table("usuarios").update({"activo": False}).eq("id", usuario_id).execute()
        return {"message": "Usuario desactivado exitosamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ============================================
# SISTEMA DE SEGURIDAD QR
# ============================================
# Agregar estos endpoints a backend/main.py ANTES de if __name__ == "__main__"

import hashlib
import secrets
from datetime import timedelta

# ==========================================
# GENERAR QR CON TOKEN ÚNICO
# ==========================================

@app.post("/api/qr/generar/{empleado_id}")
def generar_qr_token(empleado_id: int, duracion_minutos: int = 60):
    """
    Generar token QR único de un solo uso para un empleado
    El token expira después de X minutos (default 60)
    """
    try:
        supabase = get_supabase()
        
        # Verificar que el empleado existe y está activo
        empleado = supabase.table("empleados").select("*").eq("id", empleado_id).eq("activo", True).execute()
        if not empleado.data:
            raise HTTPException(status_code=404, detail="Empleado no encontrado o inactivo")
        
        empleado_data = empleado.data[0]
        
        # Generar token único y seguro
        token = secrets.token_urlsafe(32)
        
        # Generar hash de seguridad (token + empleado_id + timestamp)
        timestamp = datetime.now().isoformat()
        hash_data = f"{token}:{empleado_id}:{timestamp}"
        hash_seguridad = hashlib.sha256(hash_data.encode()).hexdigest()
        
        # Calcular fecha de expiración
        fecha_expiracion = datetime.now() + timedelta(minutes=duracion_minutos)
        
        # Guardar en base de datos
        qr_data = {
            "empleado_id": empleado_id,
            "token": token,
            "hash_seguridad": hash_seguridad,
            "fecha_expiracion": fecha_expiracion.isoformat()
        }
        
        result = supabase.table("qr_tokens").insert(qr_data).execute()
        token_record = result.data[0]
        
        # Preparar respuesta con datos del QR
        qr_payload = {
            "token": token,
            "empleado_id": empleado_id,
            "rut": empleado_data['rut'],
            "nombre": f"{empleado_data['nombre']} {empleado_data['apellido']}",
            "tipo_contrato": empleado_data['tipo_contrato'],
            "expira": fecha_expiracion.isoformat(),
            "hash": hash_seguridad[:16]  # Primeros 16 caracteres para verificación visual
        }
        
        return {
            "success": True,
            "qr_data": qr_payload,
            "qr_string": f"CLIPCONTROL:{token}:{empleado_id}",  # Formato para generar QR
            "expira_en_minutos": duracion_minutos,
            "mensaje": f"QR generado para {empleado_data['nombre']} {empleado_data['apellido']}"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar QR: {str(e)}")

# ==========================================
# VALIDAR QR ANTES DE ESCANEO
# ==========================================

@app.post("/api/qr/validar")
def validar_qr_token(token: str, periodo_id: Optional[int] = None):
    """
    Validar que un token QR es válido antes de proceder con el escaneo
    Verificaciones:
    1. Token existe
    2. No ha sido usado
    3. No ha expirado
    4. Empleado no ha retirado en este período
    """
    try:
        supabase = get_supabase()
        
        # Buscar token
        token_result = supabase.table("qr_tokens").select("*").eq("token", token).execute()
        
        if not token_result.data:
            return {
                "valido": False,
                "codigo": "TOKEN_INVALIDO",
                "mensaje": "QR inválido o no encontrado"
            }
        
        token_data = token_result.data[0]
        
        # Verificar si ya fue usado
        if token_data['usado']:
            return {
                "valido": False,
                "codigo": "TOKEN_USADO",
                "mensaje": f"Este QR ya fue utilizado el {token_data['fecha_uso']}",
                "fecha_uso": token_data['fecha_uso']
            }
        
        # Verificar expiración
        fecha_expiracion = datetime.fromisoformat(token_data['fecha_expiracion'].replace('Z', '+00:00'))
        if datetime.now(fecha_expiracion.tzinfo) > fecha_expiracion:
            return {
                "valido": False,
                "codigo": "TOKEN_EXPIRADO",
                "mensaje": "Este QR ha expirado. Genera uno nuevo.",
                "expiro": token_data['fecha_expiracion']
            }
        
        # Obtener datos del empleado
        empleado = supabase.table("empleados").select("*").eq("id", token_data['empleado_id']).execute()
        if not empleado.data or not empleado.data[0]['activo']:
            return {
                "valido": False,
                "codigo": "EMPLEADO_INACTIVO",
                "mensaje": "Empleado inactivo o no encontrado"
            }
        
        empleado_data = empleado.data[0]
        
        # Verificar si ya retiró en este período (si se proporciona periodo_id)
        if periodo_id:
            entrega_previa = supabase.table("entregas").select("id, fecha_hora").eq(
                "empleado_id", token_data['empleado_id']
            ).eq(
                "periodo_id", periodo_id
            ).eq(
                "estado", "COMPLETADO"
            ).execute()
            
            if entrega_previa.data:
                return {
                    "valido": False,
                    "codigo": "YA_RETIRO",
                    "mensaje": f"Este empleado ya retiró su beneficio el {entrega_previa.data[0]['fecha_hora']}",
                    "fecha_retiro": entrega_previa.data[0]['fecha_hora']
                }
        
        # Todo OK - QR válido
        tipo_caja = "PLANTA" if empleado_data['tipo_contrato'] == "PLANTA" else "PLAZO_FIJO"
        
        return {
            "valido": True,
            "codigo": "OK",
            "mensaje": "QR válido - Puede proceder con el registro",
            "empleado": {
                "id": empleado_data['id'],
                "rut": empleado_data['rut'],
                "nombre": empleado_data['nombre'],
                "apellido": empleado_data['apellido'],
                "nombre_completo": f"{empleado_data['nombre']} {empleado_data['apellido']}",
                "tipo_contrato": empleado_data['tipo_contrato'],
                "tipo_caja": tipo_caja,
                "sucursal_id": empleado_data['sucursal_id']
            },
            "token_id": token_data['id'],
            "tiempo_restante_minutos": int((fecha_expiracion - datetime.now(fecha_expiracion.tzinfo)).total_seconds() / 60)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al validar QR: {str(e)}")

# ==========================================
# REGISTRAR ENTREGA CON FOTO Y SEGURIDAD
# ==========================================

@app.post("/api/entregas/registrar-seguro")
async def registrar_entrega_seguro(
    qr_token_id: int = Form(...),
    empleado_id: int = Form(...),
    usuario_id: int = Form(...),
    periodo_id: Optional[int] = Form(None),
    foto: UploadFile = File(...),
    dispositivo_id: Optional[str] = Form(None),
    ip_address: Optional[str] = Form(None),
    latitud: Optional[float] = Form(None),
    longitud: Optional[float] = Form(None),
    observaciones: Optional[str] = Form("")
):
    """
    Registrar entrega con seguridad completa usando FormData
    """
    try:
        supabase = get_supabase()
        
        # Verificar token QR
        token_result = supabase.table("qr_tokens").select("*").eq("id", qr_token_id).execute()
        if not token_result.data:
            raise HTTPException(status_code=404, detail="Token QR no encontrado")
        
        token_data = token_result.data[0]
        
        # Verificar que no fue usado
        if token_data['usado']:
            raise HTTPException(status_code=400, detail="Este QR ya fue utilizado")
        
        # Verificar que no expiró
        fecha_expiracion = datetime.fromisoformat(token_data['fecha_expiracion'].replace('Z', '+00:00'))
        if datetime.now(fecha_expiracion.tzinfo) > fecha_expiracion:
            raise HTTPException(status_code=400, detail="QR expirado")
        
        # Leer foto y convertir a base64
        foto_bytes = await foto.read()
        import base64
        foto_base64 = f"data:image/jpeg;base64,{base64.b64encode(foto_bytes).decode()}"
        
        print(f"📸 Foto recibida: {len(foto_bytes)} bytes")
        
        # Calcular duración del escaneo
        fecha_generacion = datetime.fromisoformat(token_data['fecha_generacion'].replace('Z', '+00:00'))
        duracion_escaneo = int((datetime.now(fecha_generacion.tzinfo) - fecha_generacion).total_seconds())
        
        # Obtener datos del empleado para el tipo de caja
        empleado = supabase.table("empleados").select("tipo_contrato").eq("id", empleado_id).execute()
        tipo_caja = "PLANTA" if empleado.data[0]['tipo_contrato'] == "PLANTA" else "PLAZO_FIJO"
        
        # Obtener nombre del guardia
        usuario = supabase.table("usuarios").select("nombre_completo").eq("id", usuario_id).execute()
        nombre_guardia = usuario.data[0]['nombre_completo'] if usuario.data else "Guardia"
        
        # Crear registro de entrega
        entrega_data = {
            "empleado_id": empleado_id,
            "usuario_id": usuario_id,
            "periodo_id": periodo_id,
            "qr_token_id": qr_token_id,
            "foto_entrega": foto_base64,
            "dispositivo_id": dispositivo_id,
            "ip_address": ip_address,
            "latitud": latitud,
            "longitud": longitud,
            "duracion_escaneo": duracion_escaneo,
            "guardia": nombre_guardia,
            "tipo_caja": tipo_caja,
            "metodo": "QR_SEGURO",
            "estado": "COMPLETADO",
            "observaciones": observaciones
        }
        
        entrega_result = supabase.table("entregas").insert(entrega_data).execute()
        
        # Marcar token como usado
        supabase.table("qr_tokens").update({
            "usado": True,
            "fecha_uso": datetime.now().isoformat(),
            "ip_uso": ip_address,
            "dispositivo_uso": dispositivo_id
        }).eq("id", qr_token_id).execute()
        
        return {
            "success": True,
            "mensaje": "Entrega registrada exitosamente",
            "entrega_id": entrega_result.data[0]['id'],
            "duracion_escaneo_segundos": duracion_escaneo,
            "timestamp": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al registrar entrega: {str(e)}")

# ==========================================
# CONSULTAR AUDITORÍA DE SEGURIDAD
# ==========================================

@app.get("/api/seguridad/auditoria")
def get_auditoria_seguridad(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    empleado_id: Optional[int] = None,
    limit: int = 100
):
    """
    Obtener auditoría de seguridad de entregas
    Incluye toda la metadata de seguridad
    """
    try:
        supabase = get_supabase()
        query = supabase.table("v_auditoria_seguridad").select("*")
        
        if empleado_id:
            # Esta vista no tiene empleado_id directo, se filtraría por RUT
            pass
        
        result = query.limit(limit).execute()
        return result.data
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# LIMPIAR TOKENS EXPIRADOS (MANTENIMIENTO)
# ==========================================

@app.post("/api/qr/limpiar-expirados")
def limpiar_tokens_expirados():
    """
    Eliminar tokens QR expirados que no fueron usados
    Solo SUPERADMIN debería poder ejecutar esto
    """
    try:
        supabase = get_supabase()
        
        # Ejecutar función de limpieza
        result = supabase.rpc('limpiar_tokens_expirados').execute()
        
        return {
            "success": True,
            "tokens_eliminados": result.data,
            "mensaje": f"Se eliminaron {result.data} tokens expirados"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ==========================================
# GENERACIÓN MASIVA DE QR
# ==========================================
# Agregar a backend/main.py después de los otros endpoints de QR

@app.post("/api/qr/generar-masivo")
async def generar_qr_masivo(
    request: GenerarQRMasivoRequest,  # ← Body
    sucursal_id: Optional[int] = None,
    tipo_contrato: Optional[str] = None,
    duracion_minutos: int = 60
):
    """
    Generar QR para múltiples empleados a la vez
    """
    try:
        supabase = get_supabase()
        
        empleados_ids = request.empleados_ids
        
        # Si hay lista específica de empleados
        if empleados_ids and len(empleados_ids) > 0:
            print(f"🔍 Generando para IDs específicos: {empleados_ids}")
            empleados_query = supabase.table("empleados").select("*").in_("id", empleados_ids).eq("activo", True)
        else:
            print(f"🔍 Generando para TODOS (filtros: sucursal={sucursal_id}, tipo={tipo_contrato})")
            # Obtener todos los empleados activos con filtros
            empleados_query = supabase.table("empleados").select("*").eq("activo", True)
            
            if sucursal_id:
                empleados_query = empleados_query.eq("sucursal_id", sucursal_id)
            
            if tipo_contrato:
                empleados_query = empleados_query.eq("tipo_contrato", tipo_contrato.upper())
        
        empleados_result = empleados_query.execute()
        
        print(f"📊 Empleados encontrados: {len(empleados_result.data) if empleados_result.data else 0}")
        
        if not empleados_result.data or len(empleados_result.data) == 0:
            raise HTTPException(status_code=404, detail="No se encontraron empleados activos")
        
        empleados = empleados_result.data
        qr_generados = []
        errores = []
        fecha_expiracion = None
        
        # Generar QR para cada empleado
        for empleado in empleados:
            try:
                token = secrets.token_urlsafe(32)
                timestamp = datetime.now().isoformat()
                hash_data = f"{token}:{empleado['id']}:{timestamp}"
                hash_seguridad = hashlib.sha256(hash_data.encode()).hexdigest()
                fecha_expiracion = datetime.now() + timedelta(minutes=duracion_minutos)
                
                qr_data = {
                    "empleado_id": empleado['id'],
                    "token": token,
                    "hash_seguridad": hash_seguridad,
                    "fecha_expiracion": fecha_expiracion.isoformat()
                }
                
                result = supabase.table("qr_tokens").insert(qr_data).execute()
                
                qr_generados.append({
                    "empleado_id": empleado['id'],
                    "rut": empleado['rut'],
                    "nombre": f"{empleado['nombre']} {empleado['apellido']}",
                    "tipo_contrato": empleado['tipo_contrato'],
                    "token": token,
                    "qr_string": f"CLIPCONTROL:{token}:{empleado['id']}",
                    "expira": fecha_expiracion.isoformat(),
                    "hash": hash_seguridad[:16]
                })
                
                print(f"✅ QR generado para: {empleado['nombre']} {empleado['apellido']}")
                
            except Exception as e:
                print(f"❌ Error en empleado {empleado['id']}: {str(e)}")
                errores.append({
                    "empleado_id": empleado['id'],
                    "rut": empleado.get('rut', 'N/A'),
                    "nombre": f"{empleado.get('nombre', '')} {empleado.get('apellido', '')}",
                    "error": str(e)
                })
        
        print(f"🎉 Total generados: {len(qr_generados)}, Errores: {len(errores)}")
        
        return {
            "success": True,
            "total_empleados": len(empleados),
            "qr_generados": len(qr_generados),
            "qr_fallidos": len(errores),
            "duracion_minutos": duracion_minutos,
            "expira": fecha_expiracion.isoformat() if fecha_expiracion else None,
            "qr_data": qr_generados,
            "errores": errores if errores else None,
            "mensaje": f"Se generaron {len(qr_generados)} QR de {len(empleados)} empleados"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error general: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/qr/estadisticas")
def get_estadisticas_qr():
    """
    Obtener estadísticas de QR generados
    """
    try:
        supabase = get_supabase()
        
        # Total QR generados
        total = supabase.table("qr_tokens").select("id", count="exact").execute()
        
        # QR activos (no usados, no expirados)
        ahora = datetime.now().isoformat()
        activos = supabase.table("qr_tokens").select("id", count="exact").eq(
            "usado", False
        ).gt("fecha_expiracion", ahora).execute()
        
        # QR usados
        usados = supabase.table("qr_tokens").select("id", count="exact").eq("usado", True).execute()
        
        # QR expirados
        expirados = supabase.table("qr_tokens").select("id", count="exact").eq(
            "usado", False
        ).lt("fecha_expiracion", ahora).execute()
        
        return {
            "total_generados": total.count if total.count else 0,
            "activos": activos.count if activos.count else 0,
            "usados": usados.count if usados.count else 0,
            "expirados": expirados.count if expirados.count else 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/api/estadisticas/dashboard")
async def get_estadisticas_dashboard():
    try:
        # Obtener período activo
        periodo_result = supabase.table("periodos_entrega").select("*").eq("activo", True).execute()
        
        if not periodo_result.data:
            return {
                "total_empleados": 0,
                "entregas_realizadas": 0,
                "pendientes": 0,
                "porcentaje": 0,
                "total_entregas": 0,
                "periodo_activo": None
            }
        
        periodo_activo = periodo_result.data[0]
        
        # Total empleados activos
        emp_result = supabase.table("empleados").select("id", count="exact").eq("activo", True).execute()
        total_empleados = emp_result.count
        
        # Entregas DEL PERÍODO ACTIVO (directo por periodo_id)
        entregas_result = supabase.table("entregas").select("empleado_id").eq("periodo_id", periodo_activo['id']).execute()
        
        total_entregas = len(entregas_result.data)
        empleados_que_retiraron = len(set([e['empleado_id'] for e in entregas_result.data]))
        
        # Pendientes
        pendientes = total_empleados - empleados_que_retiraron
        
        # Porcentaje
        porcentaje = round((empleados_que_retiraron / total_empleados * 100), 1) if total_empleados > 0 else 0
        
        return {
            "total_empleados": total_empleados,
            "entregas_realizadas": empleados_que_retiraron,
            "pendientes": pendientes,
            "porcentaje": porcentaje,
            "total_entregas": total_entregas,
            "periodo_activo": {
                "id": periodo_activo['id'],
                "nombre": periodo_activo['nombre'],
                "fecha_inicio": periodo_activo['fecha_inicio'],
                "fecha_fin": periodo_activo['fecha_fin']
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reportes/pendientes")
async def generar_reporte_pendientes():
    try:
        # Obtener período activo
        periodo_result = supabase.table("periodos_entrega").select("*").eq("activo", True).execute()
        if not periodo_result.data:
            raise HTTPException(status_code=404, detail="No hay período activo")
        
        periodo_activo = periodo_result.data[0]
        
        # Todos los empleados activos
        empleados_result = supabase.table("v_empleados_completo").select("*").eq("activo", True).execute()
        empleados = empleados_result.data
        
        # Empleados que YA retiraron EN ESTE PERÍODO
        entregas_result = supabase.table("entregas").select("empleado_id").eq("periodo_id", periodo_activo['id']).execute()
        empleados_con_entrega = set([e['empleado_id'] for e in entregas_result.data])
        
        # Filtrar solo los pendientes
        pendientes = [emp for emp in empleados if emp['id'] not in empleados_con_entrega]
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados Pendientes"
        
        headers = ['RUT', 'Nombre Completo', 'Sucursal', 'Tipo Contrato', 'Sección', 'Email', 'Teléfono']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for emp in pendientes:
            ws.append([
                emp.get('rut', ''),
                emp.get('nombre_completo', ''),
                emp.get('sucursal', ''),
                emp.get('tipo_contrato', ''),
                emp.get('seccion', ''),
                emp.get('email', ''),
                emp.get('telefono', '')
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pendientes_{periodo_activo['nombre']}_{fecha_actual}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# ============================================
# GESTIÓN DE PERÍODOS
# ============================================

@app.get("/api/periodos")
async def get_periodos():
    try:
        result = supabase.table("periodos_entrega").select("*").order("fecha_inicio", desc=True).execute()
        return result.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/periodos/activo")
async def get_periodo_activo():
    try:
        result = supabase.table("periodos_entrega").select("*").eq("activo", True).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="No hay período activo")
        return result.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/periodos")
async def create_periodo(periodo: dict):
    try:
        # Desactivar otros períodos activos
        supabase.table("periodos_entrega").update({"activo": False}).eq("activo", True).execute()
        
        # Crear nuevo período activo
        data = {
            "nombre": periodo['nombre'],
            "fecha_inicio": periodo['fecha_inicio'],
            "fecha_fin": periodo['fecha_fin'],
            "activo": True
        }
        
        result = supabase.table("periodos_entrega").insert(data).execute()
        return result.data[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/periodos/{periodo_id}/cerrar")
async def cerrar_periodo(periodo_id: int):
    try:
        # Cerrar período
        result = supabase.table("periodos_entrega").update({"activo": False}).eq("id", periodo_id).execute()
        return {"message": "Período cerrado correctamente", "data": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/periodos/{periodo_id}/activar")
async def activar_periodo(periodo_id: int):
    try:
        # Desactivar otros períodos
        supabase.table("periodos_entrega").update({"activo": False}).eq("activo", True).execute()
        
        # Activar el seleccionado
        result = supabase.table("periodos_entrega").update({"activo": True}).eq("id", periodo_id).execute()
        return {"message": "Período activado correctamente", "data": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# ACTUALIZAR Y ELIMINAR PERÍODOS
# ============================================

@app.put("/api/periodos/{periodo_id}")
async def update_periodo(periodo_id: int, periodo: dict):
    try:
        data = {}
        if 'nombre' in periodo:
            data['nombre'] = periodo['nombre']
        if 'fecha_inicio' in periodo:
            data['fecha_inicio'] = periodo['fecha_inicio']
        if 'fecha_fin' in periodo:
            data['fecha_fin'] = periodo['fecha_fin']
        
        result = supabase.table("periodos_entrega").update(data).eq("id", periodo_id).execute()
        return {"message": "Período actualizado correctamente", "data": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/periodos/{periodo_id}")
async def delete_periodo(periodo_id: int):
    try:
        # Verificar que no sea el período activo
        periodo = supabase.table("periodos_entrega").select("*").eq("id", periodo_id).execute()
        if periodo.data and periodo.data[0].get('activo'):
            raise HTTPException(status_code=400, detail="No puedes eliminar el período activo")
        
        # Verificar que no tenga entregas
        entregas = supabase.table("entregas").select("id", count="exact").eq("periodo_id", periodo_id).execute()
        if entregas.count and entregas.count > 0:
            raise HTTPException(status_code=400, detail="No puedes eliminar un período con entregas registradas")
        
        result = supabase.table("periodos_entrega").delete().eq("id", periodo_id).execute()
        return {"message": "Período eliminado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ==========================================
# EJECUTAR
# ==========================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)

